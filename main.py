"""
JW Financial Telegram Agent v3 - 사수/부사수 구조 + 중단/인터럽트 기능
- /stop: 작업 즉시 중단
- 작업 중 메시지: 긴급 처리 후 작업 재개 or 중단 선택
"""

import os
import io
import base64
import httpx
import logging
import asyncio
from fastapi import FastAPI, Request
from contextlib import asynccontextmanager

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ──────────────────────────────────────────
# 전역 상태
# ──────────────────────────────────────────
conversation_history: list[dict] = []
MAX_HISTORY = 20
pending_task: dict | None = None       # 컨펌 대기 중인 작업
is_working: bool = False               # 팀 작업 진행 중 여부
stop_requested: bool = False           # /stop 요청 플래그
interrupt_message: str | None = None   # 작업 중 들어온 긴급 메시지

# ──────────────────────────────────────────
# 시스템 프롬프트 - JW대장
# ──────────────────────────────────────────
# DAEJANG_PROMPT — 업데이트 버전 (main.py 교체용)

DAEJANG_PROMPT = """
당신은 JW대장입니다. JW Financial Consulting의 AI 사수 에이전트입니다.

## 역할
- 진우님의 업무 지시를 받아 JW부장(부사수)에게 작업 지시
- JW부장 결과물 검토 및 품질 관리
- 최종본을 진우님에게 보고

## 운영자 컨텍스트
- 이름: 남진우 | 자산관리사/팀장 5년차 | 하나증권 연계 JW Financial
- 핵심 고객: 30대 맞벌이 직장인
- 주력 상품: 변액연금(10년납), 보장성보험 (수익 90%↑)
- 이탈 최다 구간: 2차→3차 사이 (보험 생소함)
- 콘텐츠 원칙: 비유 먼저, 인과관계 중심, 친근하되 전문적

## 스킬 라우팅 — 작업 유형 자동 판단

진우님 요청을 받으면 아래 기준으로 어떤 스킬이 필요한지 먼저 판단하고,
JW부장에게 지시할 때 해당 스킬의 프레임워크를 명시한다.

### SALES 스킬 적용 (skills/sales.md)
트리거: 상담, 미팅, 1차/2차/3차, 클로징, 계약, 제안서, 거절, 이탈, 소개요청, 레퍼럴
프레임워크:
- 대화/스크립트 → SPIN Selling 4단계 (S→P→I→N)
- 제안서/가치설계 → Hormozi 가치 방정식 + 5요소 구조
- 거절 대응 → Tracy 심리 패턴 대응
출력 원칙: 판매 언어 금지, 교육/공감 톤, 고객이 스스로 결론 내리게

### MARKETING 스킬 적용 (skills/marketing.md) ← 추가 예정
트리거: 카드뉴스, 블로그, 콘텐츠, SNS, 카카오채널, 숏폼, 유튜브, 핀사이트랩스

### FINANCE 스킬 적용 (skills/finance-consulting.md) ← 추가 예정
트리거: 포트폴리오, 시뮬레이션, 수익률, 자산배분, 매크로, 리포트

### OPERATIONS 스킬 적용 (skills/operations.md) ← 추가 예정
트리거: 자동화, 파이프라인, 노션, CRM, 프로세스 정리

## 업무 처리 방식
1. 진우님 요청 분석 → 스킬 라우팅 판단
2. JW부장에게 구체적 지시 (해당 스킬 프레임워크 명시)
3. JW부장 초안 검토:
   - 브랜드 톤앤매너 준수 여부
   - 고객 발송용: "✅ 검토 후 발송" 포함 여부
   - 금융 수치: "⚠️ 확인필요" 표기 여부
   - 실무 활용 완성도
4. 수정 필요시 구체적 재지시
5. OK면 "✅ 대장 검토 완료" 후 진우님 보고

## 검토 기준
- 판매 언어 → 교육/공감 톤으로 수정 지시
- 너무 긴 메시지 → 모바일 가독성 압축 지시
- 수치/날짜 공백 → ⚠️ 확인필요 표기 지시
- CTA 없는 고객용 콘텐츠 → CTA 추가 지시
- SPIN 순서 틀림 → 순서 교정 지시
- Hormozi 요소 누락 → 해당 요소 추가 지시
"""

# ──────────────────────────────────────────
# 시스템 프롬프트 - JW부장
# ──────────────────────────────────────────
BUJANG_PROMPT = """
당신은 JW부장입니다. JW Financial Consulting의 AI 부사수 에이전트입니다.

## 역할
- JW대장 지시를 받아 실제 작업물(초안) 생성
- 검토 피드백 반영하여 빠르게 수정

## 운영자 컨텍스트
- 이름: 남진우 | 자산관리사/팀장 5년차 | 하나증권 연계 JW Financial
- 핵심 고객: 30대 맞벌이 직장인
- 주력 상품: 변액연금(10년납), 보장성보험

## 스킬 참조
- 포트폴리오 북: 재무현황/SWOT/목표/포트폴리오/상품설명/로드맵 6섹션
- 이탈방지: D+1(보험 교육), D+2(비식별 사례), D+3(3차 미팅 프레이밍)
- 온보딩: D+0(환영), D+7(체크인), D+30(리뷰), D+90(관리미팅 제안)
- 소개 요청: 성장스토리 + 타겟 명확화 + 소개 방법 구체화
- 카카오 메시지: 3줄 이내, 이름 호명, 이모지 절제, CTA 1개

## 작업 원칙
1. 대장 지시사항 빠짐없이 반영
2. 고객 발송용: "✅ 검토 후 발송해주세요" 필수
3. 수치/날짜 모를 땐 [○○] + "⚠️ 확인필요" 명시
4. 판매 언어 금지 — 교육/공감 톤 유지
5. 수정 지시 반영 시 변경 부분 앞에 ▶ 표시
"""

# ──────────────────────────────────────────
# 작업 요청 키워드
# ──────────────────────────────────────────
TASK_KEYWORDS = [
    "만들어줘", "작성해줘", "짜줘", "초안", "써줘",
    "포트폴리오 북", "이탈방지", "온보딩", "소개 요청",
    "카드뉴스", "블로그", "리포트", "메시지", "시퀀스",
    "2차완료", "계약완료", "분석해줘", "정리해줘"
]

# ──────────────────────────────────────────
# FastAPI
# ──────────────────────────────────────────
@asynccontextmanager
async def lifespan(app: FastAPI):
    await set_webhook()
    yield

app = FastAPI(lifespan=lifespan)

TELEGRAM_TOKEN = os.environ["TELEGRAM_BOT_TOKEN"]
TELEGRAM_API = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}"
ALLOWED_CHAT_ID = int(os.environ["ALLOWED_CHAT_ID"])
ANTHROPIC_API_KEY = os.environ["ANTHROPIC_API_KEY"]
CLAUDE_MODEL = "claude-sonnet-4-5"
GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN", "")
GITHUB_REPO = os.environ.get("GITHUB_REPO", "abillion-abillion/jw-agent-workspace")
GITHUB_API_URL = "https://api.github.com"


# ──────────────────────────────────────────
# Telegram 유틸
# ──────────────────────────────────────────
async def set_webhook():
    webhook_url = os.environ.get("WEBHOOK_URL", "")
    if not webhook_url:
        return
    async with httpx.AsyncClient() as client:
        resp = await client.post(
            f"{TELEGRAM_API}/setWebhook",
            json={"url": f"{webhook_url}/webhook", "drop_pending_updates": True}
        )
        logger.info(f"Webhook 등록: {resp.json()}")


async def send_message(chat_id: int, text: str):
    chunks = [text[i:i+4000] for i in range(0, len(text), 4000)]
    async with httpx.AsyncClient() as client:
        for chunk in chunks:
            await client.post(
                f"{TELEGRAM_API}/sendMessage",
                json={"chat_id": chat_id, "text": chunk, "parse_mode": "Markdown"}
            )


async def send_typing(chat_id: int):
    async with httpx.AsyncClient() as client:
        await client.post(
            f"{TELEGRAM_API}/sendChatAction",
            json={"chat_id": chat_id, "action": "typing"}
        )


async def download_file(file_id: str) -> bytes:
    async with httpx.AsyncClient() as client:
        resp = await client.get(f"{TELEGRAM_API}/getFile?file_id={file_id}")
        file_path = resp.json()["result"]["file_path"]
        file_resp = await client.get(
            f"https://api.telegram.org/file/bot{TELEGRAM_TOKEN}/{file_path}"
        )
        return file_resp.content


# ──────────────────────────────────────────
# Claude API
# ──────────────────────────────────────────
async def call_claude(system: str, messages: list, max_tokens: int = 2000) -> str:
    headers = {
        "x-api-key": ANTHROPIC_API_KEY,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json"
    }
    async with httpx.AsyncClient(timeout=120.0) as client:
        resp = await client.post(
            "https://api.anthropic.com/v1/messages",
            headers=headers,
            json={
                "model": CLAUDE_MODEL,
                "max_tokens": max_tokens,
                "system": system,
                "messages": messages
            }
        )
        resp.raise_for_status()
        return resp.json()["content"][0]["text"]


# ──────────────────────────────────────────
# 중단 체크 헬퍼
# ──────────────────────────────────────────
async def check_stop(chat_id: int) -> bool:
    """작업 중 /stop 요청 확인. True면 중단."""
    global stop_requested, is_working
    if stop_requested:
        stop_requested = False
        is_working = False
        await send_message(chat_id, "⛔ 작업이 중단됐습니다.")
        return True
    return False


async def check_interrupt(chat_id: int) -> bool:
    """작업 중 긴급 메시지 확인. True면 처리 후 계속 여부 물어봄."""
    global interrupt_message, is_working
    if interrupt_message:
        msg = interrupt_message
        interrupt_message = None

        await send_message(chat_id, f"⚡ *긴급 처리 중...*\n`{msg}`")
        await send_typing(chat_id)

        # 대장이 긴급 질문 처리
        reply = await call_claude(
            system=DAEJANG_PROMPT,
            messages=[{"role": "user", "content": msg}],
            max_tokens=800
        )
        await send_message(chat_id, f"💬 *대장 답변*\n\n{reply}")
        await send_message(chat_id,
            "━━━━━━━━━━━━━━━━━━━━\n"
            "진행 중인 작업을 계속할까요?\n"
            "• *'계속'* — 작업 재개\n"
            "• */stop* — 작업 중단"
        )

        # 진우님 응답 대기 (최대 60초)
        for _ in range(60):
            await asyncio.sleep(1)
            if stop_requested:
                stop_requested = False
                is_working = False
                await send_message(chat_id, "⛔ 작업이 중단됐습니다.")
                return True  # 중단
            if interrupt_message == "계속":
                interrupt_message = None
                await send_message(chat_id, "▶️ 작업을 재개합니다...")
                return False  # 계속

        # 60초 무응답 → 자동 재개
        await send_message(chat_id, "⏱ 응답 없음 — 작업을 자동 재개합니다...")
        return False
    return False


# ──────────────────────────────────────────
# 사수-부사수 파이프라인
# ──────────────────────────────────────────
def is_task_request(text: str) -> bool:
    return any(k in text for k in TASK_KEYWORDS)


async def run_team_pipeline(chat_id: int, user_request: str, file_blocks: list = None):
    global is_working, stop_requested, pending_task

    is_working = True
    stop_requested = False

    await send_message(chat_id,
        "⚙️ *대장-부장 팀 작업 시작*\n"
        "작업 중단: `/stop`\n"
        "긴급 질문: 바로 메시지 보내세요"
    )

    try:
        # ── STEP 1: 대장 → 부장 지시 ──
        if await check_stop(chat_id): return
        await send_typing(chat_id)

        daejang_instruction = await call_claude(
            system=DAEJANG_PROMPT,
            messages=[{"role": "user", "content":
                f"진우님 요청:\n{user_request}\n\n"
                "JW부장에게 구체적으로 지시해주세요. "
                "'JW부장에게:'로 시작하세요."}],
            max_tokens=800
        )

        if await check_stop(chat_id): return
        await send_message(chat_id, f"📋 *대장 → 부장 지시*\n\n{daejang_instruction}")

        # ── STEP 2: 부장 초안 ──
        if await check_stop(chat_id): return
        if await check_interrupt(chat_id): return

        await send_message(chat_id, "✏️ *JW부장 초안 작성 중...*")
        await send_typing(chat_id)

        bujang_content = f"대장 지시:\n{daejang_instruction}\n\n진우님 요청:\n{user_request}"
        bujang_messages = [{"role": "user", "content": bujang_content}]

        if file_blocks:
            bujang_messages = [{"role": "user", "content":
                file_blocks + [{"type": "text", "text": bujang_content}]}]

        draft = await call_claude(
            system=BUJANG_PROMPT,
            messages=bujang_messages,
            max_tokens=2500
        )

        if await check_stop(chat_id): return
        await send_message(chat_id, f"📝 *JW부장 초안*\n\n{draft}")

        # ── STEP 3: 대장 검토 ──
        if await check_stop(chat_id): return
        if await check_interrupt(chat_id): return

        await send_message(chat_id, "🔍 *대장 검토 중...*")
        await send_typing(chat_id)

        review = await call_claude(
            system=DAEJANG_PROMPT,
            messages=[{"role": "user", "content":
                f"진우님 요청:\n{user_request}\n\n"
                f"JW부장 초안:\n{draft}\n\n"
                "검토해주세요. 수정 필요시 구체적으로 지시, "
                "OK면 '✅ 대장 검토 완료'로 마무리."}],
            max_tokens=800
        )

        if await check_stop(chat_id): return
        await send_message(chat_id, f"🔎 *대장 검토*\n\n{review}")

        # ── STEP 3-2: 수정 라운드 ──
        needs_revision = "검토 완료" not in review and "✅" not in review

        if needs_revision:
            if await check_stop(chat_id): return
            if await check_interrupt(chat_id): return

            await send_message(chat_id, "🔄 *JW부장 수정 중...*")
            await send_typing(chat_id)

            revised = await call_claude(
                system=BUJANG_PROMPT,
                messages=[
                    {"role": "user", "content": f"초안:\n{draft}"},
                    {"role": "assistant", "content": draft},
                    {"role": "user", "content":
                        f"대장 피드백:\n{review}\n\n"
                        "반영해서 수정해주세요. 변경 부분 앞에 ▶ 표시."}
                ],
                max_tokens=2500
            )

            if await check_stop(chat_id): return
            await send_message(chat_id, f"📝 *JW부장 수정본*\n\n{revised}")

            # 대장 최종 검토
            await send_typing(chat_id)
            final_review = await call_claude(
                system=DAEJANG_PROMPT,
                messages=[{"role": "user", "content":
                    f"수정본:\n{revised}\n\n최종 검토해주세요."}],
                max_tokens=400
            )

            if await check_stop(chat_id): return
            await send_message(chat_id, f"✅ *대장 최종 검토*\n\n{final_review}")
            draft = revised

        # ── STEP 4: 진우님 컨펌 요청 ──
        pending_task = {"original": user_request, "draft": draft}
        is_working = False

        await send_message(chat_id,
            "━━━━━━━━━━━━━━━━━━━━\n"
            "📌 *진우님 컨펌 요청*\n\n"
            "• *'확정'* — 그대로 사용\n"
            "• *'수정: [내용]'* — 수정 요청\n"
            "• *'다시'* — 처음부터 재작업"
        )

    except asyncio.CancelledError:
        is_working = False
        await send_message(chat_id, "⛔ 작업이 중단됐습니다.")
    except Exception as e:
        is_working = False
        logger.error(f"파이프라인 오류: {e}")
        await send_message(chat_id, f"⚠️ 작업 중 오류: {str(e)}")


# ──────────────────────────────────────────
# 진우님 컨펌 처리
# ──────────────────────────────────────────
async def handle_confirm(chat_id: int, text: str) -> bool:
    global pending_task

    if not pending_task:
        return False

    if text.strip() == "확정":
        await send_message(chat_id, "✅ 확정됐습니다! 다음 업무 말씀해주세요.")
        pending_task = None
        return True

    if text.strip() == "다시":
        original = pending_task["original"]
        pending_task = None
        await send_message(chat_id, "🔄 처음부터 다시 작업합니다...")
        asyncio.create_task(run_team_pipeline(chat_id, original))
        return True

    if text.startswith("수정:"):
        revision_note = text[3:].strip()
        original_draft = pending_task["draft"]
        await send_message(chat_id, f"🔄 *수정 반영 중...*\n요청: {revision_note}")
        await send_typing(chat_id)

        revised = await call_claude(
            system=BUJANG_PROMPT,
            messages=[
                {"role": "user", "content": f"기존 작업물:\n{original_draft}"},
                {"role": "assistant", "content": original_draft},
                {"role": "user", "content":
                    f"진우님 수정 요청:\n{revision_note}\n\n"
                    "수정해주세요. 변경 부분 앞에 ▶ 표시."}
            ],
            max_tokens=2500
        )

        review = await call_claude(
            system=DAEJANG_PROMPT,
            messages=[{"role": "user", "content":
                f"진우님 수정 요청: {revision_note}\n\n수정본:\n{revised}\n\n검토해주세요."}],
            max_tokens=400
        )

        pending_task["draft"] = revised
        await send_message(chat_id, f"📝 *수정본*\n\n{revised}")
        await send_message(chat_id, f"🔎 *대장 검토*\n\n{review}")
        await send_message(chat_id,
            "━━━━━━━━━━━━━━━━━━━━\n"
            "📌 *진우님 컨펌 요청*\n\n"
            "• *'확정'* / *'수정: [내용]'* / *'다시'*"
        )
        return True

    return False


# ──────────────────────────────────────────
# 단순 대화 (대장 직접)
# ──────────────────────────────────────────
async def ask_daejang(user_message: str, file_blocks: list = None) -> str:
    global conversation_history

    if file_blocks:
        content = file_blocks + [{"type": "text", "text": user_message or "분석해줘"}]
    else:
        content = user_message

    conversation_history.append({"role": "user", "content": content})
    if len(conversation_history) > MAX_HISTORY:
        conversation_history = conversation_history[-MAX_HISTORY:]

    reply = await call_claude(
        system=DAEJANG_PROMPT,
        messages=conversation_history,
        max_tokens=1500
    )
    conversation_history.append({"role": "assistant", "content": reply})
    return reply


# ──────────────────────────────────────────
# 파일 파싱
# ──────────────────────────────────────────
async def parse_file(file_bytes: bytes, mime_type: str, file_name: str) -> dict:
    mime = (mime_type or "").lower()
    name = (file_name or "").lower()

    if mime.startswith("image/") or name.endswith((".jpg", ".jpeg", ".png", ".webp")):
        img_mime = "image/jpeg" if name.endswith((".jpg", ".jpeg")) else "image/png"
        return {"type": "image", "source": {
            "type": "base64", "media_type": img_mime,
            "data": base64.standard_b64encode(file_bytes).decode("utf-8")}}

    if mime == "application/pdf" or name.endswith(".pdf"):
        return {"type": "document", "source": {
            "type": "base64", "media_type": "application/pdf",
            "data": base64.standard_b64encode(file_bytes).decode("utf-8")}}

    if name.endswith(".docx"):
        try:
            import docx
            doc = docx.Document(io.BytesIO(file_bytes))
            text = "\n".join([p.text for p in doc.paragraphs if p.text.strip()])
            return {"type": "text", "text": f"[Word: {file_name}]\n\n{text}"}
        except Exception as e:
            return {"type": "text", "text": f"[Word 파싱 실패: {e}]"}

    if name.endswith((".xlsx", ".xls")):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
            lines = []
            for sheet in wb.sheetnames[:3]:
                ws = wb[sheet]
                lines.append(f"[시트: {sheet}]")
                for row in ws.iter_rows(max_row=100, values_only=True):
                    row_text = "\t".join([str(c) if c is not None else "" for c in row])
                    if row_text.strip():
                        lines.append(row_text)
            return {"type": "text", "text": f"[Excel: {file_name}]\n\n" + "\n".join(lines)}
        except Exception as e:
            return {"type": "text", "text": f"[Excel 파싱 실패: {e}]"}

    try:
        text = file_bytes.decode("utf-8")
    except UnicodeDecodeError:
        try:
            text = file_bytes.decode("cp949")
        except Exception:
            return {"type": "text", "text": f"[{file_name}] 읽을 수 없는 형식"}
    return {"type": "text", "text": f"[파일: {file_name}]\n\n{text}"}


# ──────────────────────────────────────────
# 명령어 처리
# ──────────────────────────────────────────
async def handle_command(chat_id: int, text: str) -> bool:
    global stop_requested, interrupt_message, is_working
    global conversation_history, pending_task

    # /stop — 작업 중단
    if text.strip() == "/stop":
        if is_working:
            stop_requested = True
            await send_message(chat_id, "⛔ 중단 요청됨. 현재 단계 완료 후 중단됩니다...")
        else:
            await send_message(chat_id, "진행 중인 작업이 없습니다.")
        return True

    # 계속 — 인터럽트 후 작업 재개
    if text.strip() == "계속" and is_working:
        interrupt_message = "계속"
        return True

    if text.strip() == "/start":
        await send_message(chat_id,
            "👋 *JW 에이전트 팀*\n\n"
            "*작업 요청*\n"
            "• `○○ 포트폴리오 북 만들어줘`\n"
            "• `2차완료 홍길동`\n"
            "• `계약완료 홍길동`\n"
            "• `카드뉴스 초안 - [주제]`\n\n"
            "*작업 중 제어*\n"
            "• `/stop` — 즉시 중단\n"
            "• 아무 메시지 — 긴급 처리 후 재개 여부 선택\n\n"
            "*컨펌*\n"
            "• `확정` / `수정: [내용]` / `다시`\n\n"
            "/clear - 초기화"
        )
        return True

    if text.strip() == "/clear":
        conversation_history = []
        pending_task = None
        is_working = False
        stop_requested = False
        interrupt_message = None
        await send_message(chat_id, "🗑 전체 초기화 완료")
        return True

    if text.strip() == "/help":
        await send_message(chat_id,
            "*사용법*\n\n"
            "*팀 작업*\n"
            "`○○ 포트폴리오 북 만들어줘`\n"
            "`2차완료 [고객명]`\n"
            "`계약완료 [고객명]`\n"
            "`카드뉴스 초안 - [주제]`\n\n"
            "*작업 중 제어*\n"
            "`/stop` — 작업 중단\n"
            "메시지 보내기 — 긴급 처리 후 계속/중단 선택\n"
            "`계속` — 작업 재개\n\n"
            "*컨펌*\n"
            "`확정` / `수정: [내용]` / `다시`\n\n"
            "/clear — 전체 초기화"
        )
        return True

    return False


# ──────────────────────────────────────────
# 파일 메시지 처리
# ──────────────────────────────────────────
async def handle_file_message(chat_id: int, message: dict):
    await send_typing(chat_id)
    caption = message.get("caption", "이 파일을 분석해줘")

    file_id, mime_type, file_name = None, "", ""
    if "document" in message:
        doc = message["document"]
        file_id, mime_type = doc["file_id"], doc.get("mime_type", "")
        file_name = doc.get("file_name", "파일")
    elif "photo" in message:
        photo = message["photo"][-1]
        file_id, mime_type, file_name = photo["file_id"], "image/jpeg", "photo.jpg"

    if not file_id:
        await send_message(chat_id, "⚠️ 파일을 인식하지 못했습니다.")
        return

    try:
        await send_message(chat_id, f"📂 *{file_name}* 분석 중...")
        file_bytes = await download_file(file_id)
        file_block = await parse_file(file_bytes, mime_type, file_name)

        if is_task_request(caption):
            asyncio.create_task(
                run_team_pipeline(chat_id, caption, file_blocks=[file_block])
            )
        else:
            reply = await ask_daejang(caption, file_blocks=[file_block])
            await send_message(chat_id, reply)
    except Exception as e:
        logger.error(f"파일 처리 오류: {e}")
        await send_message(chat_id, f"⚠️ 파일 처리 오류: {str(e)}")


# ──────────────────────────────────────────
# Webhook
# ──────────────────────────────────────────
@app.post("/webhook")
async def webhook(request: Request):
    global interrupt_message, is_working

    data = await request.json()
    message = data.get("message")
    if not message:
        return {"ok": True}

    chat_id = message["chat"]["id"]
    if chat_id != ALLOWED_CHAT_ID:
        return {"ok": True}

    # 파일/이미지
    if "document" in message or "photo" in message:
        if is_working:
            # 작업 중 파일 → 긴급 인터럽트
            file_name = message.get("document", {}).get("file_name", "파일")
            interrupt_message = f"[파일 첨부됨: {file_name}] {message.get('caption', '')}"
            await send_message(chat_id,
                f"⚡ 작업 중 파일 수신됨.\n"
                f"현재 단계 완료 후 처리합니다.\n"
                f"지금 중단하려면 `/stop`"
            )
        else:
            await handle_file_message(chat_id, message)
        return {"ok": True}

    text = message.get("text", "").strip()
    if not text:
        return {"ok": True}

    # 명령어 (/stop, /clear 등) — 항상 최우선
    if await handle_command(chat_id, text):
        return {"ok": True}

    # 작업 진행 중 → 인터럽트로 처리
    if is_working:
        interrupt_message = text
        await send_message(chat_id,
            f"⚡ *메시지 수신됨*\n`{text}`\n\n"
            "현재 단계 완료 후 처리합니다.\n"
            "지금 중단하려면 `/stop`"
        )
        return {"ok": True}

    # 컨펌 처리
    if await handle_confirm(chat_id, text):
        return {"ok": True}

    await send_typing(chat_id)

    try:
        if is_task_request(text):
            # 비동기 태스크로 실행 (webhook 즉시 응답)
            asyncio.create_task(run_team_pipeline(chat_id, text))
        else:
            reply = await ask_daejang(text)
            await send_message(chat_id, reply)
    except Exception as e:
        logger.error(f"오류: {e}")
        await send_message(chat_id, f"⚠️ 오류: {str(e)}")

    return {"ok": True}


@app.get("/")
async def health():
    return {"status": "JW 대장-부장 팀 운영중 ✅ v3"}
